"""
bpplib_loader.py
================
BPPLIB veri setlerini yukleyen ve DPO modelleriyle test eden modul.

Desteklenen veri setleri:
  - Falkenauer U (80 ornek, N=120-1000, C=150)
  - Falkenauer T (80 ornek, N=60-501,  C=1000)
  - Scholl 1 (720 ornek, N=50-500, C=100-150)
  - Scholl 2 (480 ornek, N=50-500, C=1000)
  - Scholl 3 (10 ornek, N=200, C=100000)
  - Wascher  (17 ornek, N=27-239, C=10000)
  - Hard28   (28 ornek, N=160-200, C=1000)
"""

import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import os
import glob
import json
import openpyxl
from typing import Dict, List, Tuple, Optional


# ---------------------------------------------------------------
# BPPLIB INSTANCE PARSER
# ---------------------------------------------------------------

def parse_bpplib_instance(filepath: str) -> Dict:
    """
    BPPLIB BPP formatindaki dosyayi okur.
    Format: satir 1 = n, satir 2 = C, sonraki n satir = agirliklar
    """
    with open(filepath, 'r') as f:
        lines = [l.strip() for l in f.readlines() if l.strip()]

    n = int(lines[0])
    capacity = int(lines[1])
    weights = [int(lines[i + 2]) for i in range(n)]

    return {
        'name': os.path.basename(filepath),
        'n_items': n,
        'capacity': capacity,
        'weights': weights,
    }


def load_optimal_solutions(xlsx_path: str) -> Dict[str, Dict]:
    """
    Solutions.xlsx dosyasindan optimum sonuclari okur.
    Returns: {filename: {'lb': int, 'ub': int, 'status': str}}
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    solutions = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Ilk satir baslik
        for row in rows[1:]:
            if row[0] is None:
                continue
            name = str(row[0])
            lb = int(row[1]) if row[1] is not None else None
            ub = int(row[2]) if row[2] is not None else None

            # Status: LB == UB ise "Solved"
            status = "Solved" if lb is not None and ub is not None and lb == ub else "Open"

            solutions[name] = {
                'lb': lb,
                'ub': ub,
                'status': status,
                'dataset': sheet_name,
            }

    wb.close()
    return solutions


# ---------------------------------------------------------------
# DATASET LOADER
# ---------------------------------------------------------------

BPPLIB_BASE = os.path.join(os.path.dirname(__file__), 'bpplib_data', 'extracted')

DATASET_PATHS = {
    'falkenauer_u': os.path.join(BPPLIB_BASE, '1_Falkenauer', 'Falkenauer', 'Falkenauer U'),
    'falkenauer_t': os.path.join(BPPLIB_BASE, '1_Falkenauer', 'Falkenauer', 'Falkenauer_T'),
    'scholl_1':     os.path.join(BPPLIB_BASE, '2_Scholl', 'Scholl', 'Scholl_1'),
    'scholl_2':     os.path.join(BPPLIB_BASE, '2_Scholl', 'Scholl', 'Scholl_2'),
    'scholl_3':     os.path.join(BPPLIB_BASE, '2_Scholl', 'Scholl', 'Scholl_3'),
    'wascher':      None,  # Dinamik olarak bulunacak
    'hard28':       os.path.join(BPPLIB_BASE, '5_Hard28', 'Hard28'),
}


def _find_wascher_path():
    """Wascher dizinini bul (karakter kodlamasi sorunu)"""
    pattern = os.path.join(BPPLIB_BASE, '3_W*', 'W*')
    matches = glob.glob(pattern)
    if matches:
        return matches[0]
    return None


def load_dataset(dataset_name: str, max_items: Optional[int] = None) -> List[Dict]:
    """
    Belirtilen BPPLIB veri setini yukler.

    Args:
        dataset_name: 'falkenauer_u', 'falkenauer_t', 'scholl_1', 'scholl_2',
                      'scholl_3', 'wascher', 'hard28'
        max_items: Sadece N <= max_items olan ornekleri yukle (None = hepsi)

    Returns:
        Liste[Dict]: Her eleman {'name', 'n_items', 'capacity', 'weights'}
    """
    dataset_name = dataset_name.lower()

    if dataset_name == 'wascher':
        path = _find_wascher_path()
    else:
        path = DATASET_PATHS.get(dataset_name)

    if path is None or not os.path.exists(path):
        raise FileNotFoundError(f"Veri seti bulunamadi: {dataset_name} -> {path}")

    instances = []
    for f in sorted(glob.glob(os.path.join(path, '*.txt'))):
        inst = parse_bpplib_instance(f)
        if max_items is not None and inst['n_items'] > max_items:
            continue
        instances.append(inst)

    return instances


def load_all_datasets(max_items: Optional[int] = None) -> Dict[str, List[Dict]]:
    """Tum BPPLIB veri setlerini yukler."""
    all_data = {}
    for name in DATASET_PATHS:
        try:
            data = load_dataset(name, max_items=max_items)
            all_data[name] = data
            print(f"  {name}: {len(data)} ornek yuklendi")
        except FileNotFoundError as e:
            print(f"  {name}: ATLANILDI - {e}")
    return all_data


# ---------------------------------------------------------------
# KOLAY PROBLEM TANIMLAMASI
# ---------------------------------------------------------------

def get_easy_instances(solutions: Dict, datasets: Dict[str, List[Dict]]) -> List[Dict]:
    """
    BPPLIB'de 'Solved' olarak isaretlenen kolay problem orneklerini dondurur.
    Basvuru formu basari olcutu: "BPPLIB kutuphanesinde kolay problem olarak
    siniflandirilan problem ornekleri icin optimum sonuca ulasmak"
    """
    easy = []
    for ds_name, instances in datasets.items():
        for inst in instances:
            sol = solutions.get(inst['name'])
            if sol and sol['status'] == 'Solved':
                inst['optimal'] = sol['ub']
                inst['dataset'] = ds_name
                easy.append(inst)
    return easy


# ---------------------------------------------------------------
# MAIN: OZET BILGI
# ---------------------------------------------------------------

if __name__ == '__main__':
    print("=" * 60)
    print("BPPLIB Veri Seti Ozeti")
    print("=" * 60)

    # Tum veri setlerini yukle
    all_data = load_all_datasets()

    print(f"\nToplam: {sum(len(v) for v in all_data.values())} ornek")

    # Optimum sonuclari yukle
    xlsx_path = os.path.join(
        os.path.dirname(__file__), 'bpplib_data',
        'Instances', 'Solutions', 'Solutions.xlsx'
    )
    if os.path.exists(xlsx_path):
        solutions = load_optimal_solutions(xlsx_path)
        print(f"Optimum sonuc sayisi: {len(solutions)}")

        solved = sum(1 for v in solutions.values() if v['status'] == 'Solved')
        print(f"Cozulmus (Solved): {solved}")
        print(f"Acik (Open): {len(solutions) - solved}")

        # Kolay ornekler
        easy = get_easy_instances(solutions, all_data)
        print(f"\nKolay (Solved) ornekler: {len(easy)}")

        # Veri seti bazinda ozet
        print("\nVeri Seti Bazinda:")
        print(f"{'Veri Seti':<20} {'Toplam':>8} {'Solved':>8} {'N min':>6} {'N max':>6} {'C':>8}")
        print("-" * 60)
        for ds_name, instances in all_data.items():
            n_solved = sum(1 for inst in instances
                          if solutions.get(inst['name'], {}).get('status') == 'Solved')
            n_values = [inst['n_items'] for inst in instances]
            c_values = set(inst['capacity'] for inst in instances)
            c_str = '/'.join(str(c) for c in sorted(c_values)[:3])
            print(f"{ds_name:<20} {len(instances):>8} {n_solved:>8} "
                  f"{min(n_values):>6} {max(n_values):>6} {c_str:>8}")

    print("\nBPPLIB basariyla yuklendi.")
